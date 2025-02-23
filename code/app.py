import sys, os, sqlite3, json
from datetime import datetime, timedelta, date

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QTextEdit, QPushButton, QRadioButton, QButtonGroup, QStackedWidget,
    QTableWidget, QTableWidgetItem, QGroupBox, QInputDialog, QToolBar, QTimeEdit,
    QMessageBox, QFileDialog, QCalendarWidget, QHeaderView, QComboBox
)
from PyQt5.QtGui import QFont, QIntValidator, QDoubleValidator, QPainter, QColor, QPixmap
from PyQt5.QtCore import Qt, QTime

import openpyxl

#########################
# 데이터베이스 매니저
#########################
class DataManager:
    def __init__(self, db_path="training_log.db"):
        self.db_path = db_path
        self.conn = sqlite3.connect(self.db_path)
        self.create_table()
    
    def create_table(self):
        c = self.conn.cursor()
        # 식사 시각 정보는 제거되었습니다.
        c.execute('''CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT UNIQUE,
            running_training TEXT,
            strength_training TEXT,
            morning_heart_rate INTEGER,
            fasting_weight INTEGER,
            bedtime TEXT,
            wakeup_time TEXT,
            sleep_time REAL,
            bowel TEXT,
            meal_breakfast TEXT,
            meal_lunch TEXT,
            meal_dinner TEXT,
            meal_snack TEXT,
            comment TEXT
        )''')
        self.conn.commit()
    
    def save_log(self, log):
        c = self.conn.cursor()
        c.execute('''
            INSERT OR REPLACE INTO logs (
                date, running_training, strength_training, morning_heart_rate, fasting_weight,
                bedtime, wakeup_time, sleep_time, bowel,
                meal_breakfast, meal_lunch, meal_dinner, meal_snack, comment
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            log['date'],
            log['running_training'],
            log['strength_training'],
            log['morning_heart_rate'],
            log['fasting_weight'],
            log['bedtime'],
            log['wakeup_time'],
            log['sleep_time'],
            log['bowel'],
            log['meal_breakfast'],
            log['meal_lunch'],
            log['meal_dinner'],
            log['meal_snack'],
            log['comment']
        ))
        self.conn.commit()
    
    def get_log_by_date(self, log_date):
        c = self.conn.cursor()
        c.execute("SELECT * FROM logs WHERE date=?", (log_date,))
        return c.fetchone()
    
    def get_logs_between(self, start_date, end_date):
        c = self.conn.cursor()
        c.execute("SELECT * FROM logs WHERE date BETWEEN ? AND ? ORDER BY date", (start_date, end_date))
        return c.fetchall()
    
    def get_all_logs(self):
        c = self.conn.cursor()
        c.execute("SELECT * FROM logs ORDER BY date")
        return c.fetchall()
    
    def get_previous_weight(self, current_date):
        try:
            d = datetime.strptime(current_date, "%Y-%m-%d").date()
        except:
            return None
        prev_date = d - timedelta(days=1)
        c = self.conn.cursor()
        c.execute("SELECT fasting_weight FROM logs WHERE date=?", (prev_date.strftime("%Y-%m-%d"),))
        result = c.fetchone()
        return result[0] if result else None

#########################
# 러닝 트레이닝 한 줄 입력 위젯
#########################
class RunningTrainingLine(QWidget):
    def __init__(self, parent_layout):
        super().__init__()
        self.parent_layout = parent_layout

        self.type_combo = QComboBox()
        self.type_combo.addItems([
            "조깅", "지속주", "윈드 스프린트", "LSD", "가속주",
            "언덕 훈련", "인터벌", "TT", "크로스컨트리"
        ])
        self.type_combo.currentTextChanged.connect(self.update_input_fields)

        self.input_layout = QHBoxLayout()
        self.remove_button = QPushButton("삭제")
        self.remove_button.clicked.connect(self.remove_self)

        main_layout = QHBoxLayout()
        main_layout.addWidget(self.type_combo)
        main_layout.addLayout(self.input_layout)
        main_layout.addWidget(self.remove_button)
        self.setLayout(main_layout)

        self.update_input_fields(self.type_combo.currentText())

    def update_input_fields(self, new_type):
        while self.input_layout.count():
            item = self.input_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
        if new_type in ["언덕 훈련", "인터벌"]:
            self.m_edit = QLineEdit()
            self.m_edit.setValidator(QIntValidator(0, 10000))
            self.m_edit.setPlaceholderText("거리(m)")
            self.count_edit = QLineEdit()
            self.count_edit.setValidator(QIntValidator(0, 1000))
            self.count_edit.setPlaceholderText("횟수")
            self.input_layout.addWidget(self.m_edit)
            self.input_layout.addWidget(QLabel("×"))
            self.input_layout.addWidget(self.count_edit)
        else:
            self.value_edit = QLineEdit()
            self.value_edit.setValidator(QDoubleValidator(0.0, 1000.0, 2))
            self.value_edit.setPlaceholderText("숫자만 (km)")
            self.input_layout.addWidget(self.value_edit)
        # 평균 페이스 입력란 (분, 초)
        pace_layout = QHBoxLayout()
        pace_layout.addWidget(QLabel("평균 페이스:"))
        self.pace_min_edit = QLineEdit()
        self.pace_min_edit.setValidator(QIntValidator(0, 99))
        self.pace_min_edit.setPlaceholderText("분")
        pace_layout.addWidget(self.pace_min_edit)
        pace_layout.addWidget(QLabel("'"))
        self.pace_sec_edit = QLineEdit()
        self.pace_sec_edit.setValidator(QIntValidator(0, 59))
        self.pace_sec_edit.setPlaceholderText("초")
        pace_layout.addWidget(self.pace_sec_edit)
        pace_layout.addWidget(QLabel("''"))
        self.input_layout.addLayout(pace_layout)

    def remove_self(self):
        self.parent_layout.removeWidget(self)
        self.setParent(None)
        self.deleteLater()

    def to_dict(self):
        ttype = self.type_combo.currentText()
        if ttype in ["언덕 훈련", "인터벌"]:
            return {
                "type": ttype,
                "m": self.m_edit.text().strip(),
                "count": self.count_edit.text().strip(),
                "pace_min": self.pace_min_edit.text().strip(),
                "pace_sec": self.pace_sec_edit.text().strip()
            }
        else:
            return {
                "type": ttype,
                "value": self.value_edit.text().strip(),
                "pace_min": self.pace_min_edit.text().strip(),
                "pace_sec": self.pace_sec_edit.text().strip()
            }

#########################
# 러닝 트레이닝 그룹 위젯
#########################
class RunningTrainingWidget(QGroupBox):
    def __init__(self, title="러닝 트레이닝"):
        super().__init__(title)
        self.vlayout = QVBoxLayout()
        self.setLayout(self.vlayout)
        self.add_button = QPushButton("러닝 훈련 추가")
        self.add_button.clicked.connect(self.add_training_line)
        self.vlayout.addWidget(self.add_button)

    def add_training_line(self):
        line_widget = RunningTrainingLine(self.vlayout)
        self.vlayout.insertWidget(self.vlayout.count()-1, line_widget)

    def set_from_json(self, json_str):
        try:
            data = json.loads(json_str)
        except:
            data = []
        for i in reversed(range(self.vlayout.count())):
            item = self.vlayout.itemAt(i)
            w = item.widget()
            if isinstance(w, RunningTrainingLine):
                self.vlayout.removeWidget(w)
                w.setParent(None)
                w.deleteLater()
        self.vlayout.removeWidget(self.add_button)
        for d in data:
            line_widget = RunningTrainingLine(self.vlayout)
            self.vlayout.insertWidget(self.vlayout.count(), line_widget)
            if d.get("type") in line_widget.type_combo.model().stringList():
                idx = line_widget.type_combo.findText(d["type"])
                if idx >= 0:
                    line_widget.type_combo.setCurrentIndex(idx)
            if d.get("type") in ["언덕 훈련", "인터벌"]:
                line_widget.m_edit.setText(d.get("m", ""))
                line_widget.count_edit.setText(d.get("count", ""))
            else:
                line_widget.value_edit.setText(d.get("value", ""))
            line_widget.pace_min_edit.setText(d.get("pace_min", ""))
            line_widget.pace_sec_edit.setText(d.get("pace_sec", ""))
        self.vlayout.addWidget(self.add_button)

    def to_json(self):
        lines_data = []
        for i in range(self.vlayout.count()):
            item = self.vlayout.itemAt(i)
            w = item.widget()
            if isinstance(w, RunningTrainingLine):
                lines_data.append(w.to_dict())
        return json.dumps(lines_data, ensure_ascii=False)

#########################
# 러닝 트레이닝 파싱 함수
#########################
def parse_running_training_km(json_str):
    try:
        data = json.loads(json_str)
    except:
        data = []
    results = []
    total_km = 0.0
    for item in data:
        ttype = item.get("type", "")
        km_val = 0.0
        if ttype in ["언덕 훈련", "인터벌"]:
            try:
                m_val = float(item.get("m", "0"))
                count = float(item.get("count", "0"))
                km_val = (m_val * count) / 1000.0
            except:
                km_val = 0.0
        else:
            try:
                km_val = float(item.get("value", "0"))
            except:
                km_val = 0.0
        try:
            pace_min = int(item.get("pace_min", "0").strip()) if item.get("pace_min", "").strip() else 0
        except:
            pace_min = 0
        try:
            pace_sec = int(item.get("pace_sec", "0").strip()) if item.get("pace_sec", "").strip() else 0
        except:
            pace_sec = 0
        pace_str = f"{pace_min:02d}'{pace_sec:02d}''" if (pace_min or pace_sec) else ""
        results.append((ttype, km_val, pace_str))
        total_km += km_val
    return results, total_km

#########################
# CustomCalendarWidget (월간 뷰어)
# - 셀 상단 1/3은 날짜(일자만, 볼드; 일요일 빨간, 토요일 파란), 나머지 2/3은 로그 정보를 하단/오른쪽 정렬하여 표시
#########################
class CustomCalendarWidget(QCalendarWidget):
    def __init__(self, data_manager, parent=None):
        super().__init__(parent)
        self.data_manager = data_manager
        self.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)
    
    def paintCell(self, painter, rect, date):
        painter.save()
        dow = date.dayOfWeek()  # 1=월 ... 7=일
        date_area_height = int(rect.height() * (1/3))
        date_rect = rect.adjusted(0, 0, 0, -rect.height()+date_area_height)
        font = painter.font()
        font.setBold(True)
        painter.setFont(font)
        if dow == 7:
            painter.setPen(Qt.red)
        elif dow == 6:
            painter.setPen(Qt.blue)
        else:
            painter.setPen(Qt.black)
        day_str = date.toString("d")
        painter.drawText(date_rect, Qt.AlignCenter, day_str)
        
        log_rect = rect.adjusted(2, date_area_height, -2, -2)
        log = self.data_manager.get_log_by_date(date.toString("yyyy-MM-dd"))
        if log:
            run_json = log[2] if log[2] else "[]"
            _, total_km = parse_running_training_km(run_json)
            run_val = f"{total_km:.2f}km" if total_km > 0 else ""
            weight_val = ""
            if log[5]:
                weight_val = f"{log[5]}kg"
                prev = self.data_manager.get_previous_weight(date.toString("yyyy-MM-dd"))
                diff = (log[5] - prev) if prev is not None else 0
                if diff > 0:
                    weight_val += f" (+{diff}kg)"
                elif diff < 0:
                    weight_val += f" (-{abs(diff)}kg)"
                else:
                    weight_val += " (0kg)"
            sleep_val = f"{log[8]:.2f}시간" if log[8] else ""
            values = []
            if run_val:
                values.append(run_val)
            if weight_val:
                values.append(weight_val)
            if sleep_val:
                values.append(sleep_val)
            if values:
                painter.setPen(Qt.black)
                font2 = painter.font()
                font2.setBold(False)
                painter.setFont(font2)
                text_to_draw = "\n".join(values)
                painter.drawText(log_rect, Qt.AlignRight | Qt.AlignBottom | Qt.TextWordWrap, text_to_draw)
        painter.restore()

#########################
# DailyLogWidget (일간 입력 화면)
# - 러닝 트레이닝은 RunningTrainingWidget 사용
#########################
class DailyLogWidget(QWidget):
    def __init__(self, data_manager, switch_to_weekly_callback):
        super().__init__()
        self.data_manager = data_manager
        self.switch_to_weekly_callback = switch_to_weekly_callback
        self.initUI()
    
    def initUI(self):
        layout = QVBoxLayout()
        self.current_date = date.today().strftime("%Y-%m-%d")
        layout.addWidget(QLabel(f"날짜: {self.current_date}"))
        
        self.running_training_widget = RunningTrainingWidget("러닝 트레이닝")
        layout.addWidget(self.running_training_widget)
        
        layout.addWidget(QLabel("보강 트레이닝:"))
        self.strength_training = QTextEdit()
        self.strength_training.setPlaceholderText("보강 트레이닝 입력")
        layout.addWidget(self.strength_training)
        
        hr_layout = QHBoxLayout()
        hr_layout.addWidget(QLabel("아침 심박수:"))
        self.morning_heart_rate = QLineEdit()
        self.morning_heart_rate.setValidator(QIntValidator(0, 999))
        hr_layout.addWidget(self.morning_heart_rate)
        hr_layout.addWidget(QLabel("bpm"))
        layout.addLayout(hr_layout)
        
        weight_layout = QHBoxLayout()
        weight_layout.addWidget(QLabel("공복 몸무게:"))
        self.fasting_weight = QLineEdit()
        self.fasting_weight.setValidator(QIntValidator(0, 999))
        self.fasting_weight.textChanged.connect(self.update_weight_change)
        weight_layout.addWidget(self.fasting_weight)
        weight_layout.addWidget(QLabel("kg"))
        self.weight_change_label = QLabel("")
        weight_layout.addWidget(self.weight_change_label)
        layout.addLayout(weight_layout)
        
        time_layout = QHBoxLayout()
        time_layout.addWidget(QLabel("작일 취침 시각:"))
        self.bedtime = QTimeEdit()
        self.bedtime.setDisplayFormat("HH시 mm분")
        self.bedtime.setTime(QTime.currentTime())
        time_layout.addWidget(self.bedtime)
        time_layout.addWidget(QLabel("기상 시각:"))
        self.wakeup_time = QTimeEdit()
        self.wakeup_time.setDisplayFormat("HH시 mm분")
        self.wakeup_time.setTime(QTime.currentTime())
        time_layout.addWidget(self.wakeup_time)
        layout.addLayout(time_layout)
        
        sleep_layout = QHBoxLayout()
        sleep_layout.addWidget(QLabel("수면 시간:"))
        self.sleep_time_display = QLineEdit()
        self.sleep_time_display.setReadOnly(True)
        sleep_layout.addWidget(self.sleep_time_display)
        layout.addLayout(sleep_layout)
        
        self.bedtime.timeChanged.connect(self.compute_sleep_time)
        self.wakeup_time.timeChanged.connect(self.compute_sleep_time)
        self.compute_sleep_time()
        
        bowel_layout = QHBoxLayout()
        bowel_layout.addWidget(QLabel("배변 여부:"))
        self.bowel_o = QRadioButton("O")
        self.bowel_x = QRadioButton("X")
        self.bowel_group = QButtonGroup()
        self.bowel_group.addButton(self.bowel_o)
        self.bowel_group.addButton(self.bowel_x)
        bowel_layout.addWidget(self.bowel_o)
        bowel_layout.addWidget(self.bowel_x)
        layout.addLayout(bowel_layout)
        
        # 식사 입력: 식사 시각 정보 제거, 식사 내용만 입력
        meal_group = QGroupBox("식사")
        meal_layout = QVBoxLayout()
        
        bf_layout = QHBoxLayout()
        bf_layout.addWidget(QLabel("아침:"))
        self.meal_breakfast = QLineEdit()
        bf_layout.addWidget(self.meal_breakfast)
        meal_layout.addLayout(bf_layout)
        
        lunch_layout = QHBoxLayout()
        lunch_layout.addWidget(QLabel("점심:"))
        self.meal_lunch = QLineEdit()
        lunch_layout.addWidget(self.meal_lunch)
        meal_layout.addLayout(lunch_layout)
        
        dinner_layout = QHBoxLayout()
        dinner_layout.addWidget(QLabel("저녁:"))
        self.meal_dinner = QLineEdit()
        dinner_layout.addWidget(self.meal_dinner)
        meal_layout.addLayout(dinner_layout)
        
        snack_layout = QHBoxLayout()
        snack_layout.addWidget(QLabel("간식:"))
        self.meal_snack = QLineEdit()
        snack_layout.addWidget(self.meal_snack)
        meal_layout.addLayout(snack_layout)
        
        meal_group.setLayout(meal_layout)
        layout.addWidget(meal_group)
        
        layout.addWidget(QLabel("코멘트:"))
        self.comment = QTextEdit()
        self.comment.setPlaceholderText("코멘트 입력")
        layout.addWidget(self.comment)
        
        self.save_button = QPushButton("저장")
        self.save_button.clicked.connect(self.save_log)
        layout.addWidget(self.save_button)
        
        self.setLayout(layout)
    
    def compute_sleep_time(self):
        b_time = self.bedtime.time()
        w_time = self.wakeup_time.time()
        b_minutes = b_time.hour()*60 + b_time.minute()
        w_minutes = w_time.hour()*60 + w_time.minute()
        duration = (24*60 - b_minutes + w_minutes) if w_minutes < b_minutes else (w_minutes - b_minutes)
        self.sleep_time_display.setText(f"{duration/60.0:.2f} 시간")
    
    def update_weight_change(self):
        current_text = self.fasting_weight.text()
        if not current_text:
            self.weight_change_label.setText("")
            return
        try:
            current_weight = int(self.fasting_weight.text())
        except:
            self.weight_change_label.setText("")
            return
        prev = self.data_manager.get_previous_weight(self.current_date)
        diff = (current_weight - prev) if prev is not None else 0
        if diff > 0:
            self.weight_change_label.setText(f"(+{diff}kg)")
        elif diff < 0:
            self.weight_change_label.setText(f"(-{abs(diff)}kg)")
        else:
            self.weight_change_label.setText("(0kg)")
    
    def save_log(self):
        log = {}
        log['date'] = self.current_date
        log['running_training'] = self.running_training_widget.to_json()
        log['strength_training'] = self.strength_training.toPlainText()
        try:
            log['morning_heart_rate'] = int(self.morning_heart_rate.text())
        except:
            log['morning_heart_rate'] = 0
        try:
            log['fasting_weight'] = int(self.fasting_weight.text())
        except:
            log['fasting_weight'] = 0
        log['bedtime'] = self.bedtime.time().toString("HH:mm")
        log['wakeup_time'] = self.wakeup_time.time().toString("HH:mm")
        b_time = self.bedtime.time()
        w_time = self.wakeup_time.time()
        b_minutes = b_time.hour()*60 + b_time.minute()
        w_minutes = w_time.hour()*60 + w_time.minute()
        duration = (24*60 - b_minutes + w_minutes) if w_minutes < b_minutes else (w_minutes - b_minutes)
        log['sleep_time'] = duration/60.0
        log['bowel'] = "O" if self.bowel_o.isChecked() else ("X" if self.bowel_x.isChecked() else "")
        # 식사 시각 정보 제거
        log['meal_breakfast'] = self.meal_breakfast.text()
        log['meal_lunch'] = self.meal_lunch.text()
        log['meal_dinner'] = self.meal_dinner.text()
        log['meal_snack'] = self.meal_snack.text()
        log['comment'] = self.comment.toPlainText()
        self.data_manager.save_log(log)
        QMessageBox.information(self, "저장 완료", "훈련 일지가 저장되었습니다.")
        self.switch_to_weekly_callback()

#########################
# WeeklyLogWidget (주간 뷰어)
# - 총 13행: 0: 러닝, 1: 보강, 2: 아침 심박수, 3: 공복 몸무게, 4: 작일 취침, 5: 기상, 6: 수면, 7: 배변,
#   8: 아침, 9: 점심, 10: 저녁, 11: 간식, 12: 코멘트.
# - 식사 정보는 단순 텍스트로 표시합니다.
#########################
class WeeklyLogWidget(QWidget):
    def __init__(self, data_manager):
        super().__init__()
        self.data_manager = data_manager
        self.initUI()
    
    def initUI(self):
        layout = QVBoxLayout()
        self.table = QTableWidget(13, 7)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        today = date.today()
        monday = today - timedelta(days=today.weekday())
        dates = [(monday + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]
        self.table.setHorizontalHeaderLabels(dates)
        fields = ["러닝 트레이닝", "보강 트레이닝", "아침 심박수", "공복 몸무게",
                  "작일 취침 시각", "기상 시각", "수면 시간", "배변 여부",
                  "아침", "점심", "저녁", "간식", "코멘트"]
        self.table.setVerticalHeaderLabels(fields)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        for r in range(13):
            if r in [0, 1]:
                self.table.verticalHeader().setSectionResizeMode(r, QHeaderView.Stretch)
            else:
                self.table.verticalHeader().setSectionResizeMode(r, QHeaderView.Fixed)
                self.table.setRowHeight(r, 40)
        layout.addWidget(self.table)
        self.setLayout(layout)
        self.load_week_data()
    
    def load_week_data(self):
        for r in range(self.table.rowCount()):
            for c in range(self.table.columnCount()):
                self.table.setItem(r, c, QTableWidgetItem(""))
        today = date.today()
        monday = today - timedelta(days=today.weekday())
        dates = [(monday + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]
        for col, day_str in enumerate(dates):
            log = self.data_manager.get_log_by_date(day_str)
            if log:
                run_json = log[2] if log[2] else "[]"
                parsed_list, _ = parse_running_training_km(run_json)
                if parsed_list:
                    run_strs = []
                    for t, km, pace in parsed_list:
                        if pace:
                            run_strs.append(f"{t} - {km:.2f}km ({pace})")
                        else:
                            run_strs.append(f"{t} - {km:.2f}km")
                    run_text = "\n".join(run_strs)
                else:
                    run_text = ""
                self.table.setItem(0, col, QTableWidgetItem(run_text))
                self.table.setItem(1, col, QTableWidgetItem(log[3] or ""))
                self.table.setItem(2, col, QTableWidgetItem(f"{log[4]} bpm" if log[4] else ""))
                if log[5]:
                    prev = self.data_manager.get_previous_weight(day_str)
                    diff = (log[5] - prev) if prev is not None else 0
                    if diff > 0:
                        change_str = f"(+{diff}kg)"
                    elif diff < 0:
                        change_str = f"(-{abs(diff)}kg)"
                    else:
                        change_str = "(0kg)"
                    weight_text = f"{log[5]}kg {change_str}"
                else:
                    weight_text = ""
                self.table.setItem(3, col, QTableWidgetItem(weight_text))
                self.table.setItem(4, col, QTableWidgetItem(log[6] or ""))
                self.table.setItem(5, col, QTableWidgetItem(log[7] or ""))
                self.table.setItem(6, col, QTableWidgetItem(f"{log[8]:.2f} 시간" if log[8] else ""))
                self.table.setItem(7, col, QTableWidgetItem(log[9] or ""))
                # 식사 정보: 단순 텍스트로 표시 (식사 시각 정보 제거)
                self.table.setItem(8, col, QTableWidgetItem(log[10] or ""))
                self.table.setItem(9, col, QTableWidgetItem(log[11] or ""))
                self.table.setItem(10, col, QTableWidgetItem(log[12] or ""))
                self.table.setItem(11, col, QTableWidgetItem(log[13] or ""))
                self.table.setItem(12, col, QTableWidgetItem(log[14] or ""))
                
#########################
# MonthlyLogWidget (월간 뷰어)
# - 셀 상단 1/3은 날짜(일자만, 볼드; 일요일 빨간, 토요일 파란), 나머지 2/3은 로그 정보를 하단/오른쪽 정렬하여 표시
#########################
class MonthlyLogWidget(QWidget):
    def __init__(self, data_manager):
        super().__init__()
        self.data_manager = data_manager
        self.initUI()
    
    def initUI(self):
        layout = QVBoxLayout()
        self.calendar = CustomCalendarWidget(self.data_manager)
        layout.addWidget(self.calendar)
        self.setLayout(layout)

#########################
# 백업 데이터 불러오기 기능
#########################
def import_backup_data(data_manager):
    fileName, _ = QFileDialog.getOpenFileName(None, "백업 파일 불러오기", "", "Excel Files (*.xlsx);;All Files (*)")
    if not fileName:
        return
    try:
        wb = openpyxl.load_workbook(fileName)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            log = {
                'date': str(row[0]) if row[0] is not None else "",
                'running_training': row[1] if row[1] is not None else "",
                'strength_training': row[2] if row[2] is not None else "",
                'morning_heart_rate': int(row[3]) if row[3] is not None else 0,
                'fasting_weight': int(row[4]) if row[4] is not None else 0,
                'bedtime': row[5] if row[5] is not None else "",
                'wakeup_time': row[6] if row[6] is not None else "",
                'sleep_time': float(row[7]) if row[7] is not None else 0.0,
                'bowel': row[8] if row[8] is not None else "",
                'meal_breakfast': row[9] if row[9] is not None else "",
                'meal_lunch': row[10] if row[10] is not None else "",
                'meal_dinner': row[11] if row[11] is not None else "",
                'meal_snack': row[12] if row[12] is not None else "",
                'comment': row[13] if row[13] is not None else ""
            }
            data_manager.save_log(log)
        QMessageBox.information(None, "불러오기 완료", "백업 데이터가 성공적으로 불러와졌습니다.")
    except Exception as e:
        QMessageBox.warning(None, "오류", f"백업 데이터를 불러오는 중 오류 발생:\n{str(e)}")

#########################
# 이미지로 내보내기 기능 (주간, 월간 뷰)
#########################
def export_view_as_image(widget):
    pixmap = widget.grab()
    fileName, _ = QFileDialog.getSaveFileName(None, "이미지 파일로 내보내기", "", "PNG Files (*.png);;All Files (*)")
    if fileName:
        pixmap.save(fileName)
        QMessageBox.information(None, "내보내기 완료", f"파일이 저장되었습니다: {fileName}")

#########################
# 엑셀 내보내기 기능 (백업 전용)
#########################
def export_to_excel_backup(data_manager):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Backup"
    ws.append([
        "날짜", "러닝 트레이닝(JSON)", "보강 트레이닝", "아침 심박수", "공복 몸무게",
        "작일 취침 시각", "기상 시각", "수면 시간", "배변 여부",
        "아침", "점심", "저녁", "간식", "코멘트"
    ])
    logs = data_manager.get_all_logs()
    for log in logs:
        ws.append([
            log[1], log[2], log[3], log[4], log[5],
            log[6], log[7], log[8], log[9],
            log[10], log[11], log[12], log[13], log[14]
        ])
    fileName, _ = QFileDialog.getSaveFileName(None, "백업 파일로 내보내기", "", "Excel Files (*.xlsx);;All Files (*)")
    if fileName:
        wb.save(fileName)
        QMessageBox.information(None, "내보내기 완료", f"백업 파일이 저장되었습니다: {fileName}")

#########################
# MainWindow
#########################
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.data_manager = DataManager()
        self.initUI()
    
    def initUI(self):
        self.setWindowTitle("훈련 일지")
        self.setGeometry(100, 100, 800, 600)
        self.setFont(QFont("AppleSDGothicNeo", 10))
        
        toolbar = QToolBar()
        self.addToolBar(Qt.TopToolBarArea, toolbar)
        
        self.daily_button = QPushButton("일간")
        self.daily_button.clicked.connect(self.show_daily)
        toolbar.addWidget(self.daily_button)
        
        self.weekly_button = QPushButton("주간")
        self.weekly_button.clicked.connect(self.show_weekly)
        toolbar.addWidget(self.weekly_button)
        
        self.monthly_button = QPushButton("월간")
        self.monthly_button.clicked.connect(self.show_monthly)
        toolbar.addWidget(self.monthly_button)
        
        self.export_button = QPushButton("내보내기")
        self.export_button.clicked.connect(self.export_data)
        toolbar.addWidget(self.export_button)
        
        self.import_button = QPushButton("백업 불러오기")
        self.import_button.clicked.connect(lambda: import_backup_data(self.data_manager))
        toolbar.addWidget(self.import_button)
        
        self.stack = QStackedWidget()
        self.daily_view = DailyLogWidget(self.data_manager, self.show_weekly)
        self.weekly_view = WeeklyLogWidget(self.data_manager)
        self.monthly_view = MonthlyLogWidget(self.data_manager)
        self.stack.addWidget(self.daily_view)
        self.stack.addWidget(self.weekly_view)
        self.stack.addWidget(self.monthly_view)
        self.setCentralWidget(self.stack)
        self.show_daily()
    
    def show_daily(self):
        self.stack.setCurrentWidget(self.daily_view)
    
    def show_weekly(self):
        self.weekly_view.load_week_data()
        self.stack.setCurrentWidget(self.weekly_view)
    
    def show_monthly(self):
        self.stack.setCurrentWidget(self.monthly_view)
    
    def export_data(self):
        option, ok = QInputDialog.getItem(self, "내보내기 옵션 선택", "옵션:", ("주간", "월간", "백업"), 0, False)
        if ok and option:
            if option == "백업":
                export_to_excel_backup(self.data_manager)
            elif option == "주간":
                export_view_as_image(self.weekly_view)
            elif option == "월간":
                export_view_as_image(self.monthly_view)

#########################
# 프로그램 실행
#########################
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setFont(QFont("AppleSDGothicNeo", 10))
    mainWin = MainWindow()
    mainWin.show()
    sys.exit(app.exec_())